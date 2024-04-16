import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.chart import ScatterChart, Reference, Series

# Define the path to the directory containing the CSV files
# csv_dir = './data/bip/Piezometres'
csv_dir = os.path.join('data', 'bip', 'Piezometres')

# Get a list of all the _combine.csv files in the directory and its subdirectories
csv_files = glob.glob(os.path.join(csv_dir, '**/*_combine.csv'), recursive=True)

# get manual measures file
csv_manual_measures = os.path.join(csv_dir, 'manual_measures.csv')
manual_measures = pd.read_csv(csv_manual_measures, sep=';', decimal=',', header=0, encoding="latin1")
manual_measures['date_time'] = pd.to_datetime(manual_measures['date_time'], format='%d/%m/%Y %H:%M') # format date_time

# Create an empty dictionary to store the dataframes
dfs = {}

# Iterate over each CSV file
for csv_file in csv_files:
    if 'ZGraviere' not in csv_file:
        # Extract the folder name from the file path
        folder_name = os.path.basename(os.path.dirname(csv_file))
        
        # Read the CSV file into a pandas DataFrame
        df = pd.read_csv(csv_file, header=0)
        df['date_time'] = pd.to_datetime(df['date_time'], format='%Y-%m-%d %H:%M:%S') # format date_time
        df = df.sort_values(by='date_time')
        
        # Add the dataframe to the dictionary with the folder name as the key
        dfs[folder_name] = df

# Iterate over each dataframe in dfs
for key, df in dfs.items():
    if key != 'Baro':
        # Merge the current dataframe with the "Baro" dataframe on the "date_time" column
        merged_df = df.merge(dfs['Baro'][['date_time', 'level_m']], on='date_time', suffixes=('', '_baro'))
        
        # Calculate the corrected level by subtracting the "level_m_baro" column from the "level_m" column
        merged_df['level_corr_m'] = merged_df['level_m'] - merged_df['level_m_baro']

        # prepare manual measures
        manual_measures_df = manual_measures[manual_measures['sensor'] == key]
        manual_measures_df = manual_measures_df.sort_values(by='date_time')
        number_of_manual_measures = len(manual_measures_df)

        for i in range(number_of_manual_measures):

            # set elevation_NGF
            elevation_NGF = manual_measures_df['elevation_top_tube_ngf_m'].values[i] - manual_measures_df['level_manual_m'].values[i] - manual_measures_df['level_sensor_m'].values[i]

            # set date_time interval
            if i == 0 : # if first value
                if i+1 == number_of_manual_measures: # if only one value
                    start = merged_df['date_time'].min()
                    end = merged_df['date_time'].max()
                else : # if more than one value
                    start = merged_df['date_time'].min()
                    end = manual_measures_df['date_time'].values[i]
                filter = (merged_df['date_time'] >= start) & (merged_df['date_time'] <= end)
            else : # if not first value
                if i+1 != number_of_manual_measures: # if not last value
                    start = manual_measures_df['date_time'].values[i-1]
                    end = manual_measures_df['date_time'].values[i]
                else : # if last value
                    start = manual_measures_df['date_time'].values[i-1]
                    end = merged_df['date_time'].max()
                filter = (merged_df['date_time'] > start) & (merged_df['date_time'] <= end)
            merged_df.loc[filter, 'level_ngf_m'] = merged_df.loc[filter, 'level_corr_m'] + elevation_NGF

        # Update the original dataframe with the corrected values
        dfs[key] = merged_df.drop(columns=['level_m_baro'])

# Create a new workbook
workbook = Workbook()

# Iterate over each dataframe in dfs
for key, df in dfs.items():
    # Create a new worksheet with the key as the sheet name
    worksheet = workbook.create_sheet(title=key)
    
    # Write the DataFrame to the worksheet
    for row in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(row)

    if key != 'Baro':
        # Create a scatter chart
        chart = ScatterChart()

        # Set the title of the chart
        chart.title = key + " Level NGF"
        chart.style = 10
        chart.x_axis.title = "Date Time"
        chart.y_axis.title = "Level NGF"
        chart.width = 27
        chart.height = 10
        # Rotate the x-axis labels by 30 degrees
        # chart.x_axis.tickLblPos = "low"
        # chart.x_axis.tickLblSkip = 0
        chart.x_axis.tickLblRot = 30
        chart.legend = None
        
        # Set the x-axis data range
        x_values = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)
        y_values = Reference(worksheet, min_col=5, min_row=1, max_row=worksheet.max_row)
        
        # Create a series for the chart
        series = Series(y_values, xvalues=x_values, title_from_data=True)

        # style the series
        # Set the marker symbol to "none"
        series.marker.symbol = "none"
        # Set the line color to black and width to 1pt
        series.graphicalProperties.line.solidFill = "000000"
        series.graphicalProperties.line.width = 1

        # Add the series to the chart
        chart.series.append(series)

        # Add the chart to the worksheet
        worksheet.add_chart(chart, "H2")

# Remove the default "Sheet" worksheet
workbook.remove(workbook['Sheet'])

# Move the "Baro" worksheet to the first position in the workbook
workbook.move_sheet(workbook['Baro'], offset=-2)

# Save the workbook as an XLSX file
output_file = 'piezometry.xlsx'
workbook.save(output_file)